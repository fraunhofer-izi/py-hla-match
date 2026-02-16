import logging
from collections import defaultdict
from typing import Iterable, Union
from openpyxl import load_workbook
from contextlib import closing

from py_hla_match.exceptions import (
    MalformedHLAStringError,
    MalformedHLADataSourceError
)
from pyard.exceptions import InvalidAlleleError
from py_hla_match.hla import HLA
from py_hla_match.models import Individual, HLAPair

import pandas as pd

logger = logging.getLogger(__name__)


class HLADataSource:
    """
    Data source for HLA data. Parses HLA data from an excel or csv file.
    """

    def __init__(self, source_path: str,
                 col_idx_start: int = None,
                 col_idx_stop: int = None,
                 row_idx_start: int = 1) -> None:
        """
        Initialize the HLADataSource.

        :param source_path: Path to the excel or csv file
        :param col_idx_start: Column index to start parsing from (starting
            with first column as zero)
        :param col_idx_stop: Column index to stop parsing at (stop index
            column is included in parsing)
        :param row_idx_start: Row index to start parsing from (default is 1,
            which means the second row
        as we expect a header row)
        """
        self.source_path = source_path
        self.col_idx_start = col_idx_start
        self.col_idx_stop = col_idx_stop
        self.row_idx_start = row_idx_start

    def parse(
            self, stream: bool = False, chunk_size: int = 10000
    ) -> Union[list[Individual], Iterable[Individual]]:
        """
        Parse HLA data from an excel or csv file.

        :param stream: If True, return an iterable of individuals (default:
            False)
        :param chunk_size: Size of the chunks to read from the file (if
            streaming)
        :return: List of Individuals or an iterable of Individuals
        :raises ValueError: If the file format is not supported
        """
        if self.source_path.endswith('.xlsx'):
            return self._parse_excel(stream=stream, chunk_size=chunk_size)
        if self.source_path.endswith('.csv'):
            return self._parse_csv(stream=stream, chunk_size=chunk_size)
        raise ValueError("Unsupported file format.")

    def _parse_excel(
            self, stream: bool, chunk_size: int
    ) -> Union[list[Individual], Iterable[Individual]]:
        """
        Parse HLA data from an excel file.
        """
        if stream:
            return self._stream_excel(chunk_size=chunk_size)
        else:
            # respect row_idx_start by skipping preceding rows
            df = pd.read_excel(
                self.source_path,
                header=None,
                skiprows=self.row_idx_start
            )
            return self._parse_dataframe(df)

    def _stream_excel(self, chunk_size: int) -> Iterable[Individual]:
        """
        Stream HLA data from an Excel file in chunks using openpyxl.
        """
        # use closing to ensure wb.close() is called on generator exit
        with closing(load_workbook(self.source_path, read_only=True)) as wb:
            ws = wb.active
            # idx starting at 1
            rows = ws.iter_rows(
                min_row=self.row_idx_start + 1, values_only=True
            )
            buffer = []
            row_counter = 0  # Actual row count for tracking

            for row in rows:
                # Check for completely empty row
                if all(cell is None for cell in row):
                    continue
                if (
                    self.col_idx_start is not None and
                    self.col_idx_stop is not None
                ):
                    row = row[self.col_idx_start:self.col_idx_stop + 1]

                buffer.append((row_counter, row))
                row_counter += 1

                if len(buffer) >= chunk_size:
                    for row_idx, row_data in buffer:
                        yield self._parse_row(row_data, row_idx)
                    buffer.clear()

            # Yield remaining
            for row_idx, row_data in buffer:
                yield self._parse_row(row_data, row_idx)

    def _parse_csv(
            self, stream: bool, chunk_size: int
    ) -> Union[list[Individual], Iterable[Individual]]:
        """
        Parse HLA data from a csv file.
        """
        if stream:
            return self._stream_csv(chunk_size=chunk_size)
        else:
            # respect row_idx_start by skipping preceding rows
            df = pd.read_csv(
                self.source_path,
                header=None,
                skiprows=self.row_idx_start
            )
            return self._parse_dataframe(df)

    def _stream_csv(self, chunk_size: int) -> Iterable[Individual]:
        """
        Stream HLA data from a CSV file in chunks.
        """
        # respect row_idx_start in streaming mode
        reader = pd.read_csv(
            self.source_path,
            chunksize=chunk_size,
            header=None,
            skiprows=self.row_idx_start
        )
        for chunk in reader:
            if (
                self.col_idx_start is not None and
                self.col_idx_stop is not None
            ):
                chunk = chunk.iloc[:, self.col_idx_start:self.col_idx_stop + 1]
            for idx, row in chunk.iterrows():
                yield self._parse_row(row, idx)

    def _parse_row(self, row: Iterable[str], idx: int) -> Individual:
        """
        Parse a single row of HLA data into an Individual object.
        :param row: Iterable of HLA strings
        :param idx: Index of the row in the original data source
        :return: Individual object containing HLA pairs
        :raises MalformedHLADataSourceError: If more than two alleles are
            found for a locus
        """
        logger.debug(f"Parsing row {idx} with data: {row}")

        hla_pairs: list[HLAPair] = []
        locus_map = defaultdict(list)

        for hla_string in row:
            # skip nans (pandas), Nones (openpyxl), or empty strings
            if pd.isna(hla_string) or hla_string is None:
                continue

            # convert to string and strip whitespace
            # handles cases where data might be read as int/float or w padding
            cleaned_string = str(hla_string).strip()

            if cleaned_string == "":
                continue

            try:
                hla = HLA(cleaned_string)
                locus_map[hla.locus].append(hla)
            except MalformedHLAStringError:
                logger.error(
                    f'Encountered malformed HLA String {hla_string} in '
                    f'row {idx}. Skipping Allele.'
                )
                continue
            except InvalidAlleleError:
                logger.error(
                    f'Encountered invalid HLA Allele {hla_string} in '
                    f'row {idx}. Skipping Allele.'
                )
                continue

        for locus, alleles in locus_map.items():
            if len(alleles) > 2:
                raise MalformedHLADataSourceError(
                    f"Encountered third allele for locus {locus} in row {idx}."
                )
            if len(alleles) == 2:
                hla_pairs.append(HLAPair(hla1=alleles[0], hla2=alleles[1]))
            else:
                logger.warning(
                    f"Unpaired allele {alleles[0].allele_string} in row {idx}."
                )

        logger.debug(
            f"Successfully parsed row {idx}. Added {len(hla_pairs)} HLA pairs "
            "to individual."
        )
        return Individual(hla_data=hla_pairs)

    def _parse_dataframe(self, df: pd.DataFrame) -> list[Individual]:
        """
        Parse HLA data from a pandas DataFrame.
        """
        individuals: list[Individual] = []
        # slice the dataframe if start and end indices were given
        if self.col_idx_start is not None and self.col_idx_stop is not None:
            df = df.iloc[:, self.col_idx_start:self.col_idx_stop + 1]

        for idx, row in df.iterrows():
            # delegate to _parse_row
            individuals.append(self._parse_row(row, idx))

        return individuals
